import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, LabelFrame, Checkbutton, BooleanVar, Canvas, Scrollbar, ttk, Toplevel, Listbox, MULTIPLE # Import necessary modules
import os
import webbrowser # Import the webbrowser module
from openpyxl import load_workbook # Used for post-formatting
from openpyxl.styles import numbers # Used for post-formatting
import re # Import regex for sanitizing filenames

print("DEBUG: Script started.") # Debug print at the very beginning

class ExcelToolApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Text to Excel Converter and Excel Split Tool")

        # Allow the root window to be resizable
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # --- Create a Canvas and Scrollbar ---
        self.canvas = Canvas(root)
        self.scrollbar = Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Pack the scrollbar and canvas first
        self.scrollbar.grid(row=0, column=1, sticky='ns')
        self.canvas.grid(row=0, column=0, sticky='nsew') # Canvas expands with window

        # --- Create a Frame inside the Canvas to hold all content ---
        # This frame will hold all the stage frames
        self.content_frame = tk.Frame(self.canvas)

        # Add the content_frame to the canvas window
        self.canvas.create_window((0, 0), window=self.content_frame, anchor="nw")

        # --- Configure the canvas scroll region ---
        # This function updates the scrollable area of the canvas
        def on_frame_configure(event):
            self.canvas.configure(scrollregion=self.content_frame.bbox("all"))

        # Bind the function to the <Configure> event of the content_frame
        self.content_frame.bind("<Configure>", on_frame_configure)

        # --- Frame for Stage 1: Text to Single Excel ---
        # Pack this frame inside the content_frame
        self.frame_stage1 = LabelFrame(self.content_frame, text="Stage 1: Text to Single Excel Sheet", padx=20, pady=10)
        self.frame_stage1.pack(pady=10, padx=20, fill="x", expand=True) # Fill and expand horizontally

        tk.Label(self.frame_stage1, text="Input Text File:").grid(row=0, column=0, sticky="e", pady=5)
        self.input_text_entry = tk.Entry(self.frame_stage1, width=40)
        self.input_text_entry.grid(row=0, column=1, sticky="ew", pady=5) # Use sticky="ew" for horizontal expansion
        self.browse_text_button = tk.Button(self.frame_stage1, text="Browse...", command=self.select_input_text_file)
        self.browse_text_button.grid(row=0, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage1, text="Output Excel File (Single):").grid(row=1, column=0, sticky="e", pady=5)
        self.output_single_excel_entry = tk.Entry(self.frame_stage1, width=40)
        self.output_single_excel_entry.grid(row=1, column=1, sticky="ew", pady=5) # Use sticky="ew"
        self.browse_output_single_button = tk.Button(self.frame_stage1, text="Browse...", command=self.select_output_single_excel_file)
        self.browse_output_single_button.grid(row=1, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage1, text="Delimiter:").grid(row=2, column=0, sticky="e", pady=5)
        self.delimiter_entry = tk.Entry(self.frame_stage1, width=10)
        self.delimiter_entry.insert(0, ",") # Default delimiter
        self.delimiter_entry.grid(row=2, column=1, sticky="w", pady=5)

        # Frame to hold the two Stage 1 conversion buttons side-by-side
        stage1_button_frame = tk.Frame(self.frame_stage1)
        stage1_button_frame.grid(row=3, column=0, columnspan=3, pady=15)
        stage1_button_frame.columnconfigure(0, weight=1)
        stage1_button_frame.columnconfigure(1, weight=1)

        self.convert_single_button = tk.Button(
            stage1_button_frame, # Place inside the new frame
            text="Convert (Skip 1st/Last Row)", # Shorter text
            width=30, # Adjusted width
            bg="#0078D7",
            fg="white",
            command=self.run_stage1_conversion_skip_rows # Assign command
        )
        self.convert_single_button.grid(row=0, column=0, padx=5)

        # New button for converting without skipping rows
        self.convert_full_button = tk.Button(
            stage1_button_frame, # Place inside the new frame
            text="Convert (Keep All Rows)", # Text for the new button
            width=30, # Adjusted width
            bg="#0078D0", # Same color as the other button
            fg="white",
            command=self.run_stage1_conversion_full # Assign command
        )
        self.convert_full_button.grid(row=0, column=1, padx=5)

        # DataIQ Button (Moved to Stage 1)
        self.dataiq_button = tk.Button(
            self.frame_stage1, # Place directly in frame_stage1
            text="DataIQ",
            width=20, # Adjusted width
            bg="#4285F4", # Google Blue color
            fg="white",
            command=self.open_dataiq_url # Assign command
        )
        self.dataiq_button.grid(row=4, column=0, columnspan=3, pady=10) # Adjusted row and column span


        self.frame_stage1.columnconfigure(1, weight=1) # Allow entry fields to expand horizontally


        # --- Frame for Stage 2: Split Excel by Column Groups ---
        self.frame_stage2 = LabelFrame(self.content_frame, text="Stage 2: Split Excel by Column Groups", padx=20, pady=10)
        self.frame_stage2.pack(pady=10, padx=20, fill="x", expand=True)

        tk.Label(self.frame_stage2, text="Input Excel File:").grid(row=0, column=0, sticky="e", pady=5)
        self.input_split_excel_entry = tk.Entry(self.frame_stage2, width=40)
        self.input_split_excel_entry.grid(row=0, column=1, sticky="ew", pady=5)
        self.browse_split_excel_button = tk.Button(self.frame_stage2, text="Browse...", command=self.select_input_split_excel_file)
        self.browse_split_excel_button.grid(row=0, column=2, padx=5, pady=5)

        # Removed "Select Sheet:" label and combobox
        # Removed "Load Headers from Selected Sheet" button

        # Output Folder for Stage 2
        tk.Label(self.frame_stage2, text="Output Folder:").grid(row=1, column=0, sticky="e", pady=5) # Adjusted row
        self.output_split_folder_entry = tk.Entry(self.frame_stage2, width=40)
        self.output_split_folder_entry.grid(row=1, column=1, sticky="ew", pady=5) # Adjusted row
        self.browse_output_split_folder_button = tk.Button(self.frame_stage2, text="Browse...", command=self.select_output_split_folder)
        self.browse_output_split_folder_button.grid(row=1, column=2, padx=5, pady=5) # Adjusted row


        tk.Label(self.frame_stage2, text="Defined Column Groups:").grid(row=2, column=0, sticky="nw", pady=5) # Adjusted row
        # Listbox to display defined groups
        self.split_groups_listbox = Listbox(self.frame_stage2, height=5, width=50)
        self.split_groups_listbox.grid(row=2, column=1, sticky="nsew", pady=5) # Adjusted row
        # Scrollbar for the listbox
        self.split_groups_scrollbar = Scrollbar(self.frame_stage2, command=self.split_groups_listbox.yview)
        self.split_groups_scrollbar.grid(row=2, column=2, sticky='ns') # Adjusted row
        self.split_groups_listbox.config(yscrollcommand=self.split_groups_scrollbar.set)
        self.split_groups_listbox.bind("<<ListboxSelect>>", self.on_group_select) # Bind event


        # Buttons to manage groups
        split_group_button_frame = tk.Frame(self.frame_stage2)
        split_group_button_frame.grid(row=3, column=0, columnspan=3, pady=5) # Adjusted row
        self.add_group_button = tk.Button(split_group_button_frame, text="Add Group", state=tk.DISABLED, command=self.add_column_group) # Start disabled
        self.add_group_button.grid(row=0, column=0, padx=5)
        self.edit_group_button = tk.Button(split_group_button_frame, text="Edit Selected Group", state=tk.DISABLED, command=self.edit_selected_group) # Start disabled
        self.edit_group_button.grid(row=0, column=1, padx=5)
        self.remove_group_button = tk.Button(split_group_button_frame, text="Remove Selected Group", state=tk.DISABLED, command=self.remove_selected_group) # Start disabled
        self.remove_group_button.grid(row=0, column=2, padx=5)

        # --- Inline Frame for Defining/Editing a Group ---
        # This frame will be shown/hidden as needed
        self.group_definition_frame = LabelFrame(self.frame_stage2, text="Define/Edit Column Group", padx=10, pady=10)
        # Initially, do NOT pack this frame. It will be packed when needed.

        tk.Label(self.group_definition_frame, text="Output File Name (without .xlsx):").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.output_file_name_entry = tk.Entry(self.group_definition_frame, width=40)
        self.output_file_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew", columnspan=2)

        tk.Label(self.group_definition_frame, text="Select Columns for this Group:").grid(row=1, column=0, sticky="nw", padx=5, pady=5)

        # Frame to hold the checkboxes with a scrollbar
        self.headers_checkbox_container = tk.Frame(self.group_definition_frame)
        self.headers_checkbox_container.grid(row=1, column=1, padx=5, pady=5, sticky="nsew", columnspan=2)

        self.headers_canvas = Canvas(self.headers_checkbox_container)
        self.headers_scrollbar = Scrollbar(self.headers_checkbox_container, orient="vertical", command=self.headers_canvas.yview)
        self.headers_checkbox_frame = tk.Frame(self.headers_canvas) # Frame to hold the checkboxes

        self.headers_canvas.create_window((0, 0), window=self.headers_checkbox_frame, anchor="nw")
        self.headers_canvas.configure(yscrollcommand=self.headers_scrollbar.set)

        self.headers_scrollbar.pack(side="right", fill="y")
        self.headers_canvas.pack(side="left", fill="both", expand=True)

        # Configure scrolling for the headers checkbox frame
        def on_headers_frame_configure(event):
            self.headers_canvas.configure(scrollregion=self.headers_checkbox_frame.bbox("all"))

        self.headers_checkbox_frame.bind("<Configure>", on_headers_frame_configure)


        # Buttons for Select All / Deselect All
        select_buttons_frame = tk.Frame(self.group_definition_frame)
        select_buttons_frame.grid(row=2, column=1, columnspan=2, pady=5)
        self.select_all_button = tk.Button(select_buttons_frame, text="Select All", command=self.select_all_headers_checkboxes)
        self.select_all_button.grid(row=0, column=0, padx=5)
        self.deselect_all_button = tk.Button(select_buttons_frame, text="Deselect All", command=self.deselect_all_headers_checkboxes)
        self.deselect_all_button.grid(row=0, column=1, padx=5)


        # Save/Cancel Buttons for inline definition
        inline_button_frame = tk.Frame(self.group_definition_frame)
        inline_button_frame.grid(row=3, column=0, columnspan=3, pady=10)
        self.save_group_button = tk.Button(inline_button_frame, text="Save Group", command=self.save_column_group)
        self.save_group_button.grid(row=0, column=0, padx=5)
        self.cancel_group_button = tk.Button(inline_button_frame, text="Cancel", command=self.cancel_column_group_edit)
        self.cancel_group_button.grid(row=0, column=1, padx=5)


        self.group_definition_frame.columnconfigure(1, weight=1) # Allow entry and checkbox container to expand
        self.group_definition_frame.rowconfigure(1, weight=1) # Allow checkbox container row to expand


        self.perform_split_button = tk.Button(
            self.frame_stage2,
            text="Perform Split",
            width=50,
            bg="#28A745", # Success Green color
            fg="white",
            state=tk.DISABLED, # Start disabled until groups are defined
            command=self.perform_column_group_split # Assign command
        )
        self.perform_split_button.grid(row=4, column=0, columnspan=3, pady=15) # Adjusted row


        self.frame_stage2.columnconfigure(1, weight=1) # Allow entry fields and listbox to expand
        self.frame_stage2.rowconfigure(2, weight=1) # Allow listbox row to expand (Adjusted row)
        # Note: group_definition_frame will be packed dynamically


        # Internal instance variables to store data
        self.defined_column_groups = [] # Store defined column groups: [(output_file_name, [column_names]), ...]
        self.all_loaded_headers = [] # Store headers loaded from the selected sheet
        self.header_checkbox_vars = [] # Store BooleanVars for checkboxes
        self.editing_group_index = None # Store the index of the group being edited, None if adding new


        # --- Frame for Stage 3: Search Value in Excel Column ---
        self.frame_stage3 = LabelFrame(self.content_frame, text="Stage 3: Search Value in Excel Column", padx=20, pady=10)
        self.frame_stage3.pack(pady=10, padx=20, fill="x", expand=True)

        tk.Label(self.frame_stage3, text="Input Excel File (for Search):").grid(row=0, column=0, sticky="e", pady=5)
        self.input_search_excel_entry = tk.Entry(self.frame_stage3, width=40)
        self.input_search_excel_entry.grid(row=0, column=1, sticky="ew", pady=5)
        self.browse_search_excel_button = tk.Button(self.frame_stage3, text="Browse...", command=lambda: self.input_search_excel_entry.insert(0, filedialog.askopenfilename(title="Select Input Excel File (Stage 3)", filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])))
        self.browse_search_excel_button.grid(row=0, column=2, padx=5, pady=5)

        self.load_search_headers_button = tk.Button(self.frame_stage3, text="Load Columns", command=self.load_search_excel_columns)
        self.load_search_headers_button.grid(row=1, column=0, columnspan=3, pady=5)

        tk.Label(self.frame_stage3, text="Select Column:").grid(row=2, column=0, sticky="e", pady=5)
        self.search_column_combobox = ttk.Combobox(self.frame_stage3, width=37, state="disabled") # Start disabled
        self.search_column_combobox.grid(row=2, column=1, sticky="ew", pady=5)

        tk.Label(self.frame_stage3, text="Value to Search:").grid(row=3, column=0, sticky="e", pady=5)
        self.search_value_entry = tk.Entry(self.frame_stage3, width=40, state="disabled") # Start disabled
        self.search_value_entry.grid(row=3, column=1, sticky="ew", pady=5)

        self.search_button = tk.Button(
            self.frame_stage3,
            text="Search",
            width=20,
            bg="#17A2B8", # Info Blue color
            fg="white",
            state="disabled", # Start disabled
            command=self.perform_search # Assign command - This is the line from the traceback
        )
        self.search_button.grid(row=3, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage3, text="Search Results:").grid(row=4, column=0, sticky="nw", pady=5)
        self.search_results_text = tk.Text(self.frame_stage3, height=10, width=60, state="disabled", wrap="none") # Use Text widget for multi-line results, disable editing
        self.search_results_text.grid(row=4, column=1, columnspan=2, sticky="nsew", pady=5)

        # Add a scrollbar for the text widget
        self.search_results_scrollbar = Scrollbar(self.frame_stage3, command=self.search_results_text.yview)
        self.search_results_scrollbar.grid(row=4, column=3, sticky='ns')
        self.search_results_text.config(yscrollcommand=self.search_results_scrollbar.set)


        self.frame_stage3.columnconfigure(1, weight=1) # Allow entry fields and combobox to expand
        self.frame_stage3.rowconfigure(4, weight=1) # Allow text widget row to expand

        # --- Configure canvas scrolling when window is resized ---
        self.canvas.bind("<Configure>", self.on_canvas_configure)


    # --- Conversion and Search Methods (now part of the class) ---

    def select_input_text_file(self):
        """Opens a file dialog for selecting the input text file (Stage 1)."""
        file_path = filedialog.askopenfilename(
            title="Select Input Text File (Stage 1)",
            filetypes=[("Text Files", "*.txt *.csv"), ("All Files", "*.*")]
        )
        if file_path: # Only update if a file was selected
            self.input_text_entry.delete(0, tk.END)
            self.input_text_entry.insert(0, file_path)

    def select_output_single_excel_file(self):
        """Opens a save dialog for specifying the Stage 1 output Excel file."""
        file_path = filedialog.asksaveasfilename(
            title="Save Output Single Sheet Excel As (Stage 1)",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path: # Only update if a file was selected
            self.output_single_excel_entry.delete(0, tk.END)
            self.output_single_excel_entry.insert(0, file_path)

    def run_stage1_conversion_skip_rows(self):
        """Triggers the text to single Excel sheet conversion process (Stage 1 - Skip Rows)."""
        input_file = self.input_text_entry.get()
        output_file = self.output_single_excel_entry.get()
        delimiter = self.delimiter_entry.get()

        if not input_file:
            messagebox.showerror("Input Error", "Please select an Input Text File (Stage 1).")
            return
        if not output_file:
            messagebox.showerror("Output Error", "Please specify an Output Single Sheet Excel File (Stage 1).")
            return
        if not delimiter:
            messagebox.showerror("Input Error", "Please provide a Delimiter.")
            return
        if len(delimiter) > 1:
             messagebox.showwarning("Input Warning", "Using a multi-character delimiter might lead to unexpected results.")


        # Disable UI elements during conversion
        self.convert_single_button.config(state=tk.DISABLED)
        self.convert_full_button.config(state=tk.DISABLED) # Disable new button
        self.dataiq_button.config(state=tk.DISABLED) # Disable DataIQ button
        # Disable Stage 2 UI
        self.input_split_excel_entry.config(state=tk.DISABLED)
        self.browse_split_excel_button.config(state=tk.DISABLED)
        # self.split_sheet_name_combobox.config(state=tk.DISABLED) # Removed
        # self.load_split_headers_button.config(state=tk.DISABLED) # Removed
        self.add_group_button.config(state=tk.DISABLED)
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)
        self.split_groups_listbox.config(state=tk.DISABLED)
        self.output_split_folder_entry.config(state=tk.DISABLED)
        self.browse_output_split_folder_button.config(state=tk.DISABLED)
        # Disable Stage 3 UI
        self.input_search_excel_entry.config(state=tk.DISABLED)
        self.browse_search_excel_button.config(state=tk.DISABLED)
        self.load_search_headers_button.config(state=tk.DISABLED)
        self.search_column_combobox.config(state=tk.DISABLED)
        self.search_value_entry.config(state=tk.DISABLED)
        self.search_button.config(state=tk.DISABLED)
        self.search_results_text.config(state="normal")
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.insert(tk.END, "Stage 3 disabled during Stage 1 conversion.")
        self.search_results_text.config(state="disabled")


        self.root.update_idletasks() # Update GUI immediately

        success, msg = self.convert_text_to_excel_skip_first_last(input_file, output_file, delimiter)

        # Re-enable UI elements
        self.convert_single_button.config(state=tk.NORMAL)
        self.convert_full_button.config(state=tk.NORMAL) # Re-enable new button
        self.dataiq_button.config(state=tk.NORMAL) # Re-enable DataIQ button
        # Re-enable Stage 2 UI
        self.input_split_excel_entry.config(state=tk.NORMAL)
        self.browse_split_excel_button.config(state=tk.NORMAL)
        self.output_split_folder_entry.config(state=tk.NORMAL)
        self.browse_output_split_folder_button.config(state=tk.NORMAL)
        # split_sheet_name_combobox state is handled by load_split_excel_sheets (removed)
        # self.load_split_headers_button.config(state=tk.NORMAL) # This button was removed, this line can be removed or ignored
        # add_group_button state is handled by load_split_excel_headers
        # edit_group_button and remove_group_button state handled by listbox selection
        # perform_split_button state handled by update_groups_listbox
        self.split_groups_listbox.config(state=tk.NORMAL)
        # Re-enable Stage 3 UI
        self.input_search_excel_entry.config(state=tk.NORMAL)
        self.browse_search_excel_button.config(state=tk.NORMAL)
        self.load_search_headers_button.config(state=tk.NORMAL)
        self.search_column_combobox.config(state=tk.NORMAL) # Re-enable search column combobox
        self.search_value_entry.config(state=tk.NORMAL) # Re-enable search value entry
        self.search_button.config(state=tk.NORMAL) # Re-enable search button
        self.search_results_text.config(state="normal")
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.insert(tk.END, "Load an Excel file in Stage 3 to search.")
        self.search_results_text.config(state="disabled")
        # Re-enable sheet selection combobox in Stage 2 if sheets were loaded (Removed)
        # if self.split_sheet_name_combobox['values']:
        #      self.split_sheet_name_combobox.config(state="readonly")
        # Re-enable add group button if headers were loaded in Stage 2
        if self.all_loaded_headers:
             self.add_group_button.config(state=tk.NORMAL)
        # Re-enable perform split button if groups are defined in Stage 2
        if self.defined_column_groups:
             self.perform_split_button.config(state=tk.NORMAL)


        if success:
            messagebox.showinfo("Stage 1 Success", msg)
        else:
            messagebox.showerror("Stage 1 Failed", msg)
            # Ensure Stage 2 UI is disabled
            self.input_split_excel_entry.config(state=tk.DISABLED)
            self.browse_split_excel_button.config(state=tk.DISABLED)
            # self.split_sheet_name_combobox.config(state=tk.DISABLED) # Removed
            # self.load_split_headers_button.config(state=tk.DISABLED) # Removed
            self.add_group_button.config(state=tk.DISABLED)
            self.edit_group_button.config(state=tk.DISABLED)
            self.remove_group_button.config(state=tk.DISABLED)
            self.perform_split_button.config(state=tk.DISABLED)
            self.split_groups_listbox.config(state=tk.DISABLED)
            self.output_split_folder_entry.config(state=tk.DISABLED)
            self.browse_output_split_folder_button.config(state=tk.DISABLED)
            # Ensure Stage 3 widgets remain disabled
            self.input_search_excel_entry.config(state=tk.DISABLED)
            self.browse_search_excel_button.config(state=tk.DISABLED)
            self.load_search_headers_button.config(state=tk.DISABLED)
            self.search_column_combobox.config(state=tk.DISABLED)
            self.search_value_entry.config(state=tk.DISABLED)
            self.search_button.config(state=tk.DISABLED)
            self.search_results_text.config(state="normal")
            self.search_results_text.delete(1.0, tk.END)
            self.search_results_text.insert(tk.END, "Stage 3 disabled due to Stage 1 failure or no file selected.")
            self.search_results_text.config(state="disabled")


    def run_stage1_conversion_full(self):
        """Triggers the text to single Excel sheet conversion process (Stage 1 - Full Convert)."""
        input_file = self.input_text_entry.get()
        output_file = self.output_single_excel_entry.get()
        delimiter = self.delimiter_entry.get()

        if not input_file:
            messagebox.showerror("Input Error", "Please select an Input Text File (Stage 1).")
            return
        if not output_file:
            messagebox.showerror("Output Error", "Please specify an Output Single Sheet Excel File (Stage 1).")
            return
        if not delimiter:
            messagebox.showerror("Input Error", "Please provide a Delimiter.")
            return
        if len(delimiter) > 1:
             messagebox.showwarning("Input Warning", "Using a multi-character delimiter might lead to unexpected results.")


        # Disable UI elements during conversion
        self.convert_single_button.config(state=tk.DISABLED)
        self.convert_full_button.config(state=tk.DISABLED) # Disable new button
        self.dataiq_button.config(state=tk.DISABLED) # Disable DataIQ button
        # Disable Stage 2 UI
        self.input_split_excel_entry.config(state=tk.DISABLED)
        self.browse_split_excel_button.config(state=tk.DISABLED)
        # self.split_sheet_name_combobox.config(state=tk.DISABLED) # Removed
        # self.load_split_headers_button.config(state=tk.DISABLED) # Removed
        self.add_group_button.config(state=tk.DISABLED)
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)
        self.split_groups_listbox.config(state=tk.DISABLED)
        self.output_split_folder_entry.config(state=tk.DISABLED)
        self.browse_output_split_folder_button.config(state=tk.DISABLED)
        # Disable Stage 3 UI
        self.input_search_excel_entry.config(state=tk.DISABLED)
        self.browse_search_excel_button.config(state=tk.DISABLED)
        self.load_search_headers_button.config(state=tk.DISABLED)
        self.search_column_combobox.config(state=tk.DISABLED)
        self.search_value_entry.config(state=tk.DISABLED)
        self.search_button.config(state=tk.DISABLED)
        self.search_results_text.config(state="normal")
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.insert(tk.END, "Stage 3 disabled during Stage 1 conversion.")
        self.search_results_text.config(state="disabled")


        self.root.update_idletasks() # Update GUI immediately

        success, msg = self.convert_text_to_excel_full(input_file, output_file, delimiter) # Call the new function

        # Re-enable UI elements
        self.convert_single_button.config(state=tk.NORMAL)
        self.convert_full_button.config(state=tk.NORMAL) # Re-enable new button
        self.dataiq_button.config(state=tk.NORMAL) # Re-enable DataIQ button
        # Re-enable Stage 2 UI
        self.input_split_excel_entry.config(state=tk.NORMAL)
        self.browse_split_excel_button.config(state=tk.NORMAL)
        self.output_split_folder_entry.config(state=tk.NORMAL)
        self.browse_output_split_folder_button.config(state=tk.NORMAL)
        # split_sheet_name_combobox state is handled by load_split_excel_sheets (removed)
        # self.load_split_headers_button.config(state=tk.NORMAL) # This button was removed, this line can be removed or ignored
        # add_group_button state is handled by load_split_excel_headers
        # edit_group_button and remove_group_button state handled by listbox selection
        # perform_split_button state handled by update_groups_listbox
        self.split_groups_listbox.config(state=tk.NORMAL)
        # Re-enable Stage 3 UI
        self.input_search_excel_entry.config(state=tk.NORMAL)
        self.browse_search_excel_button.config(state=tk.NORMAL)
        self.load_search_headers_button.config(state=tk.NORMAL)
        self.search_column_combobox.config(state=tk.NORMAL) # Re-enable search column combobox
        self.search_value_entry.config(state=tk.NORMAL) # Re-enable search value entry
        self.search_button.config(state=tk.NORMAL) # Re-enable search button
        self.search_results_text.config(state="normal")
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.insert(tk.END, "Load an Excel file in Stage 3 to search.")
        self.search_results_text.config(state="disabled")
        # Re-enable sheet selection combobox in Stage 2 if sheets were loaded (Removed)
        # if self.split_sheet_name_combobox['values']:
        #      self.split_sheet_name_combobox.config(state="readonly")
        # Re-enable add group button if headers were loaded in Stage 2
        if self.all_loaded_headers:
             self.add_group_button.config(state=tk.NORMAL)
        # Re-enable perform split button if groups are defined in Stage 2
        if self.defined_column_groups:
             self.perform_split_button.config(state=tk.NORMAL)


        if success:
            messagebox.showinfo("Stage 1 Success", msg)
        else:
            messagebox.showerror("Stage 1 Failed", msg)
            # Ensure Stage 2 UI is disabled
            self.input_split_excel_entry.config(state=tk.DISABLED)
            self.browse_split_excel_button.config(state=tk.DISABLED)
            # self.split_sheet_name_combobox.config(state=tk.DISABLED) # Removed
            # self.load_split_headers_button.config(state=tk.DISABLED) # Removed
            self.add_group_button.config(state=tk.DISABLED)
            self.edit_group_button.config(state=tk.DISABLED)
            self.remove_group_button.config(state=tk.DISABLED)
            self.perform_split_button.config(state=tk.DISABLED)
            self.split_groups_listbox.config(state=tk.DISABLED)
            self.output_split_folder_entry.config(state=tk.DISABLED)
            self.browse_output_split_folder_button.config(state=tk.DISABLED)
            # Ensure Stage 3 widgets remain disabled
            self.input_search_excel_entry.config(state=tk.DISABLED)
            self.browse_search_excel_button.config(state=tk.DISABLED)
            self.load_search_headers_button.config(state=tk.DISABLED)
            self.search_column_combobox.config(state=tk.DISABLED)
            self.search_value_entry.config(state=tk.DISABLED)
            self.search_button.config(state=tk.DISABLED)
            self.search_results_text.config(state="normal")
            self.search_results_text.delete(1.0, tk.END)
            self.search_results_text.insert(tk.END, "Stage 3 disabled due to Stage 1 failure or no file selected.")
            self.search_results_text.config(state="disabled")


    # --- Stage 2: Split Excel by Column Groups Methods ---

    def select_input_split_excel_file(self):
        """Opens a file dialog for selecting the input Excel file (Stage 2)."""
        file_path = filedialog.askopenfilename(
            title="Select Input Excel File (Stage 2)",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path: # Only update if a file was selected
            self.input_split_excel_entry.delete(0, tk.END)
            self.input_split_excel_entry.insert(0, file_path)
            # Directly load headers from the first sheet after selecting the file
            print(f"DEBUG: File selected in Stage 2: {file_path}") # Debug print
            self.load_split_excel_headers()
            # Clear defined groups when a new file is selected
            self.clear_defined_groups()
            # Hide the inline definition frame
            self.hide_group_definition_frame()
            print(f"DEBUG: State of self.all_loaded_headers after select_input_split_excel_file completes: {self.all_loaded_headers}") # Debug print


    # Removed load_split_excel_sheets method

    def load_split_excel_headers(self): # Removed event parameter
        """
        Loads column headers from the first sheet of the input Excel file (Stage 2),
        stores them internally, creates checkboxes, and enables group management buttons.
        Triggered by selecting a file.
        """
        print("DEBUG: Inside load_split_excel_headers.") # Debug print
        input_excel_file = self.input_split_excel_entry.get()

        # Clear previously loaded headers and disable group management/split buttons
        self.all_loaded_headers = [] # <-- This clears the headers at the START of the function
        self.add_group_button.config(state=tk.DISABLED)
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)
        self.clear_defined_groups() # Clear any previously defined groups
        self.clear_header_checkboxes() # Clear existing checkboxes
        self.hide_group_definition_frame() # Hide the inline definition frame


        if not input_excel_file:
            print("DEBUG: No input file selected in load_split_excel_headers.") # Debug print
            return

        if not os.path.exists(input_excel_file):
             print(f"DEBUG: File not found in load_split_excel_headers: {input_excel_file}") # Debug print
             messagebox.showwarning("File Not Found", f"Input Excel file not found: {input_excel_file}")
             return

        try:
            # Use pandas to read headers from the first sheet (sheet_name=0)
            print(f"DEBUG: Attempting to read headers from {input_excel_file}, sheet 0.") # Debug print
            headers_df = pd.read_excel(input_excel_file, sheet_name=0, nrows=0)
            headers = headers_df.columns.tolist()

            if headers:
                self.all_loaded_headers = headers # Store headers internally HERE
                print(f"DEBUG: Headers loaded successfully inside load_split_excel_headers: {self.all_loaded_headers}") # Debug print
                # Create checkboxes for each header
                self.create_header_checkboxes(headers)

                # Enable group management buttons
                self.add_group_button.config(state=tk.NORMAL)
                print("DEBUG: Add Group button enabled.") # Debug print
                # Edit and Remove buttons are enabled when a group is selected in the listbox

            else:
                 print("DEBUG: No headers found in the first sheet.") # Debug print
                 messagebox.showwarning("No Headers Found", f"Could not detect headers in the first sheet of Excel file: {input_excel_file}.\nCheck if the first row contains headers.")
                 # Buttons remain disabled

            print(f"DEBUG: State of self.all_loaded_headers at end of load_split_excel_headers: {self.all_loaded_headers}") # Debug print before return


        except FileNotFoundError: # This might still happen if path changes after check
            print(f"DEBUG: FileNotFoundError in load_split_excel_headers: {input_excel_file}") # Debug print
            messagebox.showerror("File Not Found", f"Input Excel file not found: {input_excel_file}")
        except Exception as e:
            print(f"DEBUG: Exception in load_split_excel_headers: {e}") # Debug print
            messagebox.showerror("Error Loading Excel Headers", str(e))

    def create_header_checkboxes(self, headers):
        """Creates checkboxes for each header in the headers_checkbox_frame."""
        # Clear previous checkboxes and their variables
        self.clear_header_checkboxes()

        self.header_checkbox_vars = []
        for i, header in enumerate(headers):
            var = BooleanVar()
            cb = Checkbutton(self.headers_checkbox_frame, text=header, variable=var, anchor="w")
            cb.grid(row=i, column=0, sticky="w")
            self.header_checkbox_vars.append(var)

        # Update the scroll region after adding checkboxes
        self.headers_checkbox_frame.update_idletasks()
        self.on_headers_frame_configure(None) # Trigger scroll region update


    def clear_header_checkboxes(self):
        """Removes all header checkboxes from the headers_checkbox_frame."""
        for widget in self.headers_checkbox_frame.winfo_children():
            widget.destroy()
        self.header_checkbox_vars = [] # Clear the list of variables
        self.all_loaded_headers = [] # Clear loaded headers


    def select_all_headers_checkboxes(self):
        """Selects all header checkboxes."""
        for var in self.header_checkbox_vars:
            var.set(True)

    def deselect_all_headers_checkboxes(self):
        """Deselects all header checkboxes."""
        for var in self.header_checkbox_vars:
            var.set(False)


    def clear_defined_groups(self):
        """Clears the internal list of defined groups and updates the listbox display."""
        self.defined_column_groups = []
        self.update_groups_listbox()
        # Disable edit, remove, and perform split buttons
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)


    def update_groups_listbox(self):
        """Updates the display in the split_groups_listbox."""
        self.split_groups_listbox.delete(0, tk.END) # Clear current items
        if not self.defined_column_groups:
             self.split_groups_listbox.insert(tk.END, "No column groups defined yet.")
             # Disable edit, remove, and perform split buttons
             self.edit_group_button.config(state=tk.DISABLED)
             self.remove_group_button.config(state=tk.DISABLED)
             self.perform_split_button.config(state=tk.DISABLED)
             return

        for i, (output_file_name, columns) in enumerate(self.defined_column_groups):
            display_text = f"Group {i+1}: {', '.join(columns)} -> {output_file_name}.xlsx" # Display file name
            self.split_groups_listbox.insert(tk.END, display_text)

        # Enable perform split button if there are groups defined
        self.perform_split_button.config(state=tk.NORMAL)


    def show_group_definition_frame(self):
        """Shows the inline group definition/edit frame."""
        self.group_definition_frame.grid(row=5, column=0, columnspan=3, sticky="ew", pady=10) # Adjusted row
        # Disable other Stage 2 buttons while defining/editing
        self.add_group_button.config(state=tk.DISABLED)
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)
        self.split_groups_listbox.config(state=tk.DISABLED)
        self.output_split_folder_entry.config(state=tk.DISABLED)
        self.browse_output_split_folder_button.config(state=tk.DISABLED)


    def hide_group_definition_frame(self):
        """Hides the inline group definition/edit frame."""
        self.group_definition_frame.grid_forget()
        # Re-enable Stage 2 buttons (state will be managed by other functions)
        self.add_group_button.config(state=tk.NORMAL)
        # edit_group_button and remove_group_button state handled by listbox selection
        # perform_split_button state handled by update_groups_listbox
        self.split_groups_listbox.config(state=tk.NORMAL)
        self.output_split_folder_entry.config(state=tk.NORMAL)
        self.browse_output_split_folder_button.config(state=tk.NORMAL)
        # Ensure edit/remove buttons are correctly enabled/disabled after hiding
        self.on_group_select(None) # Trigger the listbox selection handler


    def add_column_group(self):
        """Prepares the inline frame to define a new column group."""
        print("DEBUG: Add Group button clicked.") # Debug print
        print(f"DEBUG: State of self.all_loaded_headers at start of add_column_group: {self.all_loaded_headers}") # Debug print
        if not self.all_loaded_headers:
            print("DEBUG: Headers not loaded when Add Group clicked.") # Debug print
            messagebox.showwarning("Headers Not Loaded", "Please load headers from the Excel file first.")
            return

        print("DEBUG: Headers are loaded. Proceeding to add group.") # Debug print
        self.editing_group_index = None # Indicate adding a new group

        self.output_file_name_entry.delete(0, tk.END) # Clear the entry
        self.deselect_all_headers_checkboxes() # Deselect all checkboxes

        self.group_definition_frame.config(text="Define New Column Group")
        self.show_group_definition_frame()


    def edit_selected_group(self):
        """Prepares the inline frame to edit the currently selected column group."""
        selected_indices = self.split_groups_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Selection Error", "Please select a column group from the list to edit.")
            return

        self.editing_group_index = selected_indices[0] # Store the index of the group being edited

        output_file_name, columns = self.defined_column_groups[self.editing_group_index]

        self.output_file_name_entry.delete(0, tk.END)
        self.output_file_name_entry.insert(0, output_file_name)

        # Deselect all checkboxes first
        self.deselect_all_headers_checkboxes()
        # Select the checkboxes for the columns in the selected group
        for col in columns:
            try:
                index = self.all_loaded_headers.index(col)
                self.header_checkbox_vars[index].set(True)
            except ValueError:
                print(f"Warning: Column '{col}' not found in current headers when editing.")


        self.group_definition_frame.config(text=f"Edit Column Group {self.editing_group_index + 1}")
        self.show_group_definition_frame()


    def save_column_group(self):
        """Saves the currently defined/edited column group."""
        output_file_name = self.output_file_name_entry.get().strip()
        selected_columns = [self.all_loaded_headers[i] for i, var in enumerate(self.header_checkbox_vars) if var.get()]

        if not output_file_name:
            messagebox.showwarning("Input Error", "Please specify an output file name.")
            return
        # Sanitize the file name to remove potentially invalid characters
        output_file_name = re.sub(r'[^\w\s.-]', '', output_file_name)
        output_file_name = output_file_name.replace(' ', '_') # Replace spaces with underscores for safety


        if not selected_columns:
            messagebox.showwarning("Selection Error", "Please select at least one column for this group.")
            return

        if self.editing_group_index is None: # Adding a new group
            self.defined_column_groups.append((output_file_name, selected_columns))
        else: # Editing an existing group
            self.defined_column_groups[self.editing_group_index] = (output_file_name, selected_columns)

        self.update_groups_listbox() # Update the main listbox display
        self.hide_group_definition_frame() # Hide the definition frame
        self.editing_group_index = None # Reset editing state


    def cancel_column_group_edit(self):
        """Cancels the current group definition/edit."""
        self.hide_group_definition_frame() # Hide the definition frame
        self.editing_group_index = None # Reset editing state


    def remove_selected_group(self):
        """Removes the currently selected column group."""
        selected_indices = self.split_groups_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Selection Error", "Please select a column group from the list to remove.")
            return
        # Assuming single selection is desired for removal
        group_index = selected_indices[0]
        del self.defined_column_groups[group_index]
        self.update_groups_listbox() # Update the main listbox display


    def select_output_split_folder(self):
        """Opens a directory dialog for selecting the output folder for Stage 2."""
        folder_path = filedialog.askdirectory(
            title="Select Output Folder for Split Files (Stage 2)"
        )
        if folder_path: # Only update if a folder was selected
            self.output_split_folder_entry.delete(0, tk.END)
            self.output_split_folder_entry.insert(0, folder_path)


    def perform_column_group_split(self):
        """
        Splits the input Excel file into multiple files based on the defined column groups.
        """
        input_excel_file = self.input_split_excel_entry.get()
        # Removed selected_sheet variable
        output_folder = self.output_split_folder_entry.get()


        if not input_excel_file:
            messagebox.showerror("Input Error", "Please select an Input Excel File (Stage 2).")
            return
        # Removed check for selected_sheet
        if not output_folder:
            messagebox.showerror("Input Error", "Please specify an Output Folder (Stage 2).")
            return
        if not self.defined_column_groups:
            messagebox.showwarning("No Groups Defined", "Please define at least one column group to perform the split.")
            return

        if not os.path.exists(output_folder):
            try:
                os.makedirs(output_folder)
            except Exception as e:
                messagebox.showerror("Folder Creation Error", f"Could not create output folder: {e}")
                return


        # Prepare lists for the split function (using file names now)
        output_file_names = [group[0] for group in self.defined_column_groups]
        column_groups_list = [group[1] for group in self.defined_column_groups]

        # Disable UI elements during splitting
        self.add_group_button.config(state=tk.DISABLED)
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)
        self.split_groups_listbox.config(state=tk.DISABLED)
        self.output_split_folder_entry.config(state=tk.DISABLED)
        self.browse_output_split_folder_button.config(state=tk.DISABLED)
        # Disable Stage 1 and Stage 3 UI
        self.convert_single_button.config(state=tk.DISABLED)
        self.convert_full_button.config(state=tk.DISABLED)
        self.dataiq_button.config(state=tk.DISABLED)
        self.input_search_excel_entry.config(state=tk.DISABLED)
        self.browse_search_excel_button.config(state=tk.DISABLED)
        self.load_search_headers_button.config(state=tk.DISABLED)
        self.search_column_combobox.config(state=tk.DISABLED)
        self.search_value_entry.config(state=tk.DISABLED)
        self.search_button.config(state=tk.DISABLED)
        self.search_results_text.config(state="normal")
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.insert(tk.END, "Stage 3 disabled during Stage 2 split.")
        self.search_results_text.config(state="disabled")


        self.root.update_idletasks() # Update GUI immediately

        try:
            # Read the input Excel file from the first sheet (sheet_name=0)
            df = pd.read_excel(input_excel_file, sheet_name=0, header=0, dtype=str)

            # Perform the split and save each group
            split_count = 0
            for output_file_name, columns_to_include in zip(output_file_names, column_groups_list):
                # Construct the full output file path
                output_file_path = os.path.join(output_folder, f"{output_file_name}.xlsx")
                try:
                    # Check if all columns for this group exist in the DataFrame
                    missing_cols = [col for col in columns_to_include if col not in df.columns]
                    if missing_cols:
                         messagebox.showwarning("Missing Columns", f"Skipping group for '{output_file_name}.xlsx' due to missing columns in the first sheet: {', '.join(missing_cols)}")
                         continue # Skip this group and proceed to the next

                    df_subset = df[columns_to_include]  # Select the specified columns

                    df_subset.to_excel(output_file_path, index=False)  # Save to a new Excel file

                    # Post-process with openpyxl for text formatting
                    if os.path.exists(output_file_path) and os.path.getsize(output_file_path) > 100: # Check if file was created with content
                         wb = load_workbook(output_file_path)
                         ws = wb.active
                         text_fmt = '@'
                         for row in ws.iter_rows():
                             for cell in row:
                                 cell.number_format = text_fmt
                         wb.save(output_file_path)
                    else:
                         print(f"Warning: No data or error saving subset for group '{output_file_name}.xlsx'. File might be empty or not created.") # Optional: log warning


                    split_count += 1
                except Exception as save_error:
                    print(f"Error saving group to '{output_file_path}': {save_error}") # Log individual save errors
                    messagebox.showwarning("Save Error", f"Could not save group to '{output_file_path}': {save_error}") # Show warning for individual save errors


            if split_count > 0:
                 messagebox.showinfo("Split Success", f"Successfully split Excel file into {split_count} files in folder: {output_folder}")
            else:
                 messagebox.showwarning("Split Completed", f"Split operation completed, but no files were successfully created in folder: {output_folder}")


        except FileNotFoundError:
            messagebox.showerror("File Not Found", f"Input Excel file not found at {input_excel_file}")
        except Exception as e:
            messagebox.showerror("Split Failed", str(e))

        finally:
            # Re-enable UI elements
            self.convert_single_button.config(state=tk.NORMAL)
            self.convert_full_button.config(state=tk.NORMAL)
            self.dataiq_button.config(state=tk.NORMAL)
            self.input_search_excel_entry.config(state=tk.NORMAL)
            self.browse_search_excel_button.config(state=tk.NORMAL)
            self.load_search_headers_button.config(state=tk.NORMAL)
            self.search_column_combobox.config(state=tk.NORMAL) # Re-enable search column combobox
            self.search_value_entry.config(state=tk.NORMAL) # Re-enable search value entry
            self.search_button.config(state=tk.NORMAL) # Re-enable search button
            self.search_results_text.config(state="normal")
            self.search_results_text.delete(1.0, tk.END)
            self.search_results_text.insert(tk.END, "Load an Excel file in Stage 3 to search.")
            self.search_results_text.config(state="disabled")
            # Re-enable Stage 2 UI based on loaded headers and defined groups
            self.input_split_excel_entry.config(state=tk.NORMAL)
            self.browse_split_excel_button.config(state=tk.NORMAL)
            self.output_split_folder_entry.config(state=tk.NORMAL)
            self.browse_output_split_folder_button.config(state=tk.NORMAL)
            # split_sheet_name_combobox state is handled by load_split_excel_sheets (removed)
            # self.load_split_headers_button.config(state=tk.NORMAL) # This button was removed, this line can be removed or ignored
            if self.all_loaded_headers:
                 self.add_group_button.config(state=tk.NORMAL)
            if self.defined_column_groups:
                 self.perform_split_button.config(state=tk.NORMAL)
                 # Edit and Remove buttons are enabled by listbox selection event
            self.split_groups_listbox.config(state=tk.NORMAL)


    # --- Stage 3: Search Value in Excel Column Methods ---

    def load_search_excel_columns(self):
        """
        Loads column headers from the selected input Excel file (Stage 3)
        into the search_column_combobox.
        Enables search-related widgets if headers are found.
        """
        input_excel_file = self.input_search_excel_entry.get()

        # Clear previous options and disable widgets
        self.search_column_combobox.set('') # Clear current selection
        self.search_column_combobox['values'] = [] # Clear dropdown options
        self.search_column_combobox.config(state="disabled")
        self.search_value_entry.delete(0, tk.END)
        self.search_value_entry.config(state="disabled")
        self.search_button.config(state="disabled")
        self.search_results_text.config(state="normal")
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.insert(tk.END, "Load an Excel file to see columns.")
        self.search_results_text.config(state="disabled")


        if not input_excel_file:
            messagebox.showwarning("Input Error", "Please select an Input Excel File for Search (Stage 3).")
            return

        if not os.path.exists(input_excel_file):
             messagebox.showwarning("File Not Found", f"Input Excel file not found: {input_excel_file}")
             return

        try:
            # Use pandas to read headers from the Excel file
            headers_df = pd.read_excel(input_excel_file, nrows=0)
            headers = headers_df.columns.tolist()

            if headers:
                # Populate the combobox with headers
                self.search_column_combobox['values'] = headers
                self.search_column_combobox.config(state="readonly") # Make it selectable but not editable
                self.search_results_text.config(state="normal")
                self.search_results_text.delete(1.0, tk.END)
                self.search_results_text.insert(tk.END, f"Columns loaded. Select a column and enter a value to search.")
                self.search_results_text.config(state="disabled")

                # Enable search value entry and search button
                self.search_value_entry.config(state="normal")
                self.search_button.config(state="normal")

            else:
                 messagebox.showwarning("No Headers Found", f"Could not detect headers in Excel file: {input_excel_file}.\nCheck if the first row contains headers.")
                 self.search_results_text.config(state="normal")
                 self.search_results_text.delete(1.0, tk.END)
                 self.search_results_text.insert(tk.END, "No headers found in the selected file.")
                 self.search_results_text.config(state="disabled")


        except FileNotFoundError: # This might still happen if path changes after check
            messagebox.showerror("File Not Found", f"Input Excel file not found: {input_excel_file}")
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, "Error: File not found.")
            self.search_results_text.config(state="disabled")
        except Exception as e:
            messagebox.showerror("Error Loading Excel Headers", str(e))
            self.search_results_text.config(state="normal")
            self.search_results_text.delete(1.0, tk.END)
            self.search_results_text.insert(tk.END, f"An error occurred during the search: {e}")
            self.search_results_text.config(state="disabled")


    def perform_search(self):
        """
        Reads the selected Excel file, searches for a value in the selected column,
        and displays matching rows in the results text widget.
        """
        input_excel_file = self.input_search_excel_entry.get()
        selected_column = self.search_column_combobox.get()
        search_value = self.search_value_entry.get()

        # Clear previous results
        self.search_results_text.config(state="normal")
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.config(state="disabled")


        if not input_excel_file:
            messagebox.showwarning("Input Error", "Please select an Input Excel File for Search (Stage 3).")
            return
        if not selected_column:
            messagebox.showwarning("Selection Error", "Please select a column to search in.")
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, "Please select a column.")
            self.search_results_text.config(state="disabled")
            return
        if not search_value:
            messagebox.showwarning("Input Error", "Please enter a value to search for.")
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, "Please enter a value to search.")
            self.search_results_text.config(state="disabled")
            return

        if not os.path.exists(input_excel_file):
             messagebox.showwarning("File Not Found", f"Input Excel file not found: {input_excel_file}")
             self.search_results_text.config(state="normal")
             self.search_results_text.insert(tk.END, "Error: File not found.")
             self.search_results_text.config(state="disabled")
             return


        try:
            # Read the entire Excel sheet as string to ensure consistent search
            df = pd.read_excel(input_excel_file, dtype=str)

            # Ensure the selected column exists after reading
            if selected_column not in df.columns:
                 messagebox.showerror("Column Error", f"Selected column '{selected_column}' not found in the Excel file.")
                 self.search_results_text.config(state="normal")
                 self.search_results_text.insert(tk.END, f"Error: Column '{selected_column}' not found.")
                 self.search_results_text.config(state="disabled")
                 return

            # Perform the search: filter rows where the selected column contains the search value
            # Use .astype(str) and .str.contains() for case-insensitive substring search
            # .fillna('') is important to handle potential NaN values in the column before searching
            # regex=False ensures that the search value is treated as a literal string, not a regular expression
            matching_rows_df = df[df[selected_column].astype(str).str.contains(search_value, case=False, na=False, regex=False)]


            # Display results
            self.search_results_text.config(state="normal")
            self.search_results_text.delete(1.0, tk.END) # Clear previous results

            if not matching_rows_df.empty:
                # Display the matching rows (including headers)
                # Convert the DataFrame to a string representation
                results_string = matching_rows_df.to_string(index=False)
                self.search_results_text.insert(tk.END, results_string)
            else:
                self.search_results_text.insert(tk.END, f"No results found for '{search_value}' in column '{selected_column}'.")

            self.search_results_text.config(state="disabled") # Disable editing

        except Exception as e:
            messagebox.showerror("Search Error", str(e))
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, f"An error occurred during the search: {e}")
            self.search_results_text.config(state="disabled")


    def open_dataiq_url(self):
        """Opens the DataIQ URL in the default web browser."""
        dataiq_url = "https://www.example.com/dataiq" # *** Replace with your actual DataIQ URL ***
        try:
            webbrowser.open(dataiq_url)
        except Exception as e:
            messagebox.showerror("Error Opening URL", f"Could not open the DataIQ URL: {e}")


    def on_group_select(self, event):
        """Handles listbox selection to enable/disable edit/remove buttons."""
        # Check if the group definition frame is currently visible
        if self.group_definition_frame.winfo_ismapped():
            # If the frame is visible, keep edit/remove disabled regardless of selection
            self.edit_group_button.config(state=tk.DISABLED)
            self.remove_group_button.config(state=tk.DISABLED)
        else:
            # Otherwise, enable/disable based on listbox selection
            if self.split_groups_listbox.curselection():
                self.edit_group_button.config(state=tk.NORMAL)
                self.remove_group_button.config(state=tk.NORMAL)
            else:
                self.edit_group_button.config(state=tk.DISABLED)
                self.remove_group_button.config(state=tk.DISABLED)


    def on_headers_frame_configure(self, event):
        """Configures the scroll region for the headers checkbox canvas."""
        self.headers_canvas.configure(scrollregion=self.headers_checkbox_frame.bbox("all"))

    def on_canvas_configure(self, event):
        """Configures the scroll region for the main canvas when the window is resized."""
        self.canvas.itemconfig(self.canvas.find_withtag("all"), width=self.canvas.winfo_width())
        # Also update the headers canvas scroll region if visible
        if self.headers_canvas.winfo_ismapped():
             self.headers_canvas.configure(scrollregion=self.headers_checkbox_frame.bbox("all"))


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelToolApp(root)
    root.mainloop()
