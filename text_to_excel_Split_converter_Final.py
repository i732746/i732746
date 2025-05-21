import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, LabelFrame, Checkbutton, BooleanVar, Canvas, Scrollbar, ttk, Listbox
import os
import webbrowser
from openpyxl import load_workbook
import re
import platform

class ExcelToolApp:
    def __init__(self, root):
        # Dynamically adjust size based on system configuration
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        width = min(950, int(screen_width * 0.8))
        height = min(800, int(screen_height * 0.8))
        root.geometry(f"{width}x{height}")

        self.root = root
        self.root.title("Text to Excel Converter and Excel Split Tool")
        if platform.system() == "Windows":
            default_font = ("Segoe UI", 10)
        else:
            default_font = ("Arial", 11)
        self.root.option_add("*Font", default_font)

        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # Outer frame, canvas, and scrollbar for full app scrollability
        self.outer_frame = tk.Frame(root)
        self.outer_frame.grid(row=0, column=0, sticky="nsew")
        self.outer_frame.grid_rowconfigure(0, weight=1)
        self.outer_frame.grid_columnconfigure(0, weight=1)

        self.outer_canvas = Canvas(self.outer_frame, borderwidth=0)
        self.vscrollbar = Scrollbar(self.outer_frame, orient="vertical", command=self.outer_canvas.yview)
        self.hscrollbar = Scrollbar(self.outer_frame, orient="horizontal", command=self.outer_canvas.xview)
        self.outer_canvas.configure(yscrollcommand=self.vscrollbar.set, xscrollcommand=self.hscrollbar.set)
        self.vscrollbar.grid(row=0, column=1, sticky="ns")
        self.hscrollbar.grid(row=1, column=0, sticky="ew")
        self.outer_canvas.grid(row=0, column=0, sticky="nsew")

        self.content_frame = tk.Frame(self.outer_canvas)
        self.content_frame_id = self.outer_canvas.create_window((0, 0), window=self.content_frame, anchor="nw")

        def on_frame_configure(event):
            self.outer_canvas.configure(scrollregion=self.outer_canvas.bbox("all"))
        def on_canvas_configure(event):
            canvas_width = event.width
            self.outer_canvas.itemconfig(self.content_frame_id, width=canvas_width)
        self.content_frame.bind("<Configure>", on_frame_configure)
        self.outer_canvas.bind("<Configure>", on_canvas_configure)

        # Mousewheel scrolling for the whole frame
        def _on_mousewheel(event):
            self.outer_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.outer_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # --- Stage 1: Text to Single Excel Sheet ---
        self.frame_stage1 = LabelFrame(self.content_frame, text="Stage 1: Text to Single Excel Sheet", padx=20, pady=10)
        self.frame_stage1.pack(pady=10, padx=20, fill="x", expand=True)

        tk.Label(self.frame_stage1, text="Input Text File:").grid(row=0, column=0, sticky="e", pady=5)
        self.input_text_entry = tk.Entry(self.frame_stage1, width=40)
        self.input_text_entry.grid(row=0, column=1, sticky="ew", pady=5)
        self.browse_text_button = tk.Button(self.frame_stage1, text="Browse...", command=self.select_input_text_file)
        self.browse_text_button.grid(row=0, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage1, text="Output Excel File (Single):").grid(row=1, column=0, sticky="e", pady=5)
        self.output_single_excel_entry = tk.Entry(self.frame_stage1, width=40)
        self.output_single_excel_entry.grid(row=1, column=1, sticky="ew", pady=5)
        self.browse_output_single_button = tk.Button(self.frame_stage1, text="Browse...", command=self.select_output_single_excel_file)
        self.browse_output_single_button.grid(row=1, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage1, text="Delimiter:").grid(row=2, column=0, sticky="e", pady=5)
        self.delimiter_entry = tk.Entry(self.frame_stage1, width=10)
        self.delimiter_entry.insert(0, ",")
        self.delimiter_entry.grid(row=2, column=1, sticky="w", pady=5)

        stage1_button_frame = tk.Frame(self.frame_stage1)
        stage1_button_frame.grid(row=3, column=0, columnspan=3, pady=15)
        stage1_button_frame.columnconfigure(0, weight=1)
        stage1_button_frame.columnconfigure(1, weight=1)

        self.convert_single_button = tk.Button(
            stage1_button_frame,
            text="Convert (Skip 1st/Last Row)",
            width=30,
            bg="#0078D7",
            fg="white",
            command=self.run_stage1_conversion_skip_rows
        )
        self.convert_single_button.grid(row=0, column=0, padx=5)

        self.convert_full_button = tk.Button(
            stage1_button_frame,
            text="Convert (Keep All Rows)",
            width=30,
            bg="#0078D0",
            fg="white",
            command=self.run_stage1_conversion_full
        )
        self.convert_full_button.grid(row=0, column=1, padx=5)

        self.frame_stage1.columnconfigure(1, weight=1)

        # --- Stage 2: Split Excel by Column Groups ---
        self.frame_stage2 = LabelFrame(self.content_frame, text="Stage 2: Split Excel by Column Groups", padx=20, pady=10)
        self.frame_stage2.pack(pady=10, padx=20, fill="x", expand=True)

        tk.Label(self.frame_stage2, text="Input Excel File:").grid(row=0, column=0, sticky="e", pady=5)
        self.input_split_excel_entry = tk.Entry(self.frame_stage2, width=40)
        self.input_split_excel_entry.grid(row=0, column=1, sticky="ew", pady=5)
        self.browse_split_excel_button = tk.Button(self.frame_stage2, text="Browse...", command=self.select_input_split_excel_file)
        self.browse_split_excel_button.grid(row=0, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage2, text="Output Folder:").grid(row=1, column=0, sticky="e", pady=5)
        self.output_split_folder_entry = tk.Entry(self.frame_stage2, width=40)
        self.output_split_folder_entry.grid(row=1, column=1, sticky="ew", pady=5)
        self.browse_output_split_folder_button = tk.Button(self.frame_stage2, text="Browse...", command=self.select_output_split_folder)
        self.browse_output_split_folder_button.grid(row=1, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage2, text="Defined Column Groups:").grid(row=2, column=0, sticky="nw", pady=5)
        self.split_groups_listbox_frame = tk.Frame(self.frame_stage2)
        self.split_groups_listbox_frame.grid(row=2, column=1, sticky="nsew", pady=5)
        self.split_groups_listbox = Listbox(self.split_groups_listbox_frame, height=5, width=50)
        self.split_groups_listbox.pack(side="left", fill="both", expand=True)
        self.split_groups_scrollbar = Scrollbar(self.split_groups_listbox_frame, command=self.split_groups_listbox.yview)
        self.split_groups_scrollbar.pack(side="right", fill="y")
        self.split_groups_listbox.config(yscrollcommand=self.split_groups_scrollbar.set)
        self.split_groups_listbox.bind("<<ListboxSelect>>", self.on_group_select)

        split_group_button_frame = tk.Frame(self.frame_stage2)
        split_group_button_frame.grid(row=3, column=0, columnspan=3, pady=5)
        self.add_group_button = tk.Button(split_group_button_frame, text="Add Group", state=tk.DISABLED, command=self.add_column_group)
        self.add_group_button.grid(row=0, column=0, padx=5)
        self.edit_group_button = tk.Button(split_group_button_frame, text="Edit Selected Group", state=tk.DISABLED, command=self.edit_selected_group)
        self.edit_group_button.grid(row=0, column=1, padx=5)
        self.remove_group_button = tk.Button(split_group_button_frame, text="Remove Selected Group", state=tk.DISABLED, command=self.remove_selected_group)
        self.remove_group_button.grid(row=0, column=2, padx=5)

        self.group_definition_frame = LabelFrame(self.frame_stage2, text="Define/Edit Column Group", padx=10, pady=10)

        tk.Label(self.group_definition_frame, text="Output File Name (without .xlsx):").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.output_file_name_entry = tk.Entry(self.group_definition_frame, width=40)
        self.output_file_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew", columnspan=2)

        tk.Label(self.group_definition_frame, text="Select Columns for this Group:").grid(row=1, column=0, sticky="nw", padx=5, pady=5)
        self.headers_checkbox_container = tk.Frame(self.group_definition_frame)
        self.headers_checkbox_container.grid(row=1, column=1, padx=5, pady=5, sticky="nsew", columnspan=2)

        self.headers_canvas = Canvas(self.headers_checkbox_container)
        self.headers_scrollbar = Scrollbar(self.headers_checkbox_container, orient="vertical", command=self.headers_canvas.yview)
        self.headers_checkbox_frame = tk.Frame(self.headers_canvas)
        self.headers_canvas.create_window((0, 0), window=self.headers_checkbox_frame, anchor="nw")
        self.headers_canvas.configure(yscrollcommand=self.headers_scrollbar.set)
        self.headers_scrollbar.pack(side="right", fill="y")
        self.headers_canvas.pack(side="left", fill="both", expand=True)
        def on_headers_frame_configure(event):
            self.headers_canvas.configure(scrollregion=self.headers_checkbox_frame.bbox("all"))
        self.headers_checkbox_frame.bind("<Configure>", on_headers_frame_configure)

        select_buttons_frame = tk.Frame(self.group_definition_frame)
        select_buttons_frame.grid(row=2, column=1, columnspan=2, pady=5)
        self.select_all_button = tk.Button(select_buttons_frame, text="Select All", command=self.select_all_headers_checkboxes)
        self.select_all_button.grid(row=0, column=0, padx=5)
        self.deselect_all_button = tk.Button(select_buttons_frame, text="Deselect All", command=self.deselect_all_headers_checkboxes)
        self.deselect_all_button.grid(row=0, column=1, padx=5)

        inline_button_frame = tk.Frame(self.group_definition_frame)
        inline_button_frame.grid(row=3, column=0, columnspan=3, pady=10)
        self.save_group_button = tk.Button(inline_button_frame, text="Save Group", command=self.save_column_group)
        self.save_group_button.grid(row=0, column=0, padx=5)
        self.cancel_group_button = tk.Button(inline_button_frame, text="Cancel", command=self.cancel_column_group_edit)
        self.cancel_group_button.grid(row=0, column=1, padx=5)

        self.group_definition_frame.columnconfigure(1, weight=1)
        self.group_definition_frame.rowconfigure(1, weight=1)

        self.perform_split_button = tk.Button(
            self.frame_stage2,
            text="Perform Split",
            width=50,
            bg="#28A745",
            fg="white",
            state=tk.DISABLED,
            command=self.perform_column_group_split
        )
        self.perform_split_button.grid(row=4, column=0, columnspan=3, pady=15)

        self.dataiq_button_stage2 = tk.Button(
            self.frame_stage2,
            text="DataIQ",
            width=20,
            bg="#4285F4",
            fg="white",
            command=self.open_dataiq_url
        )
        self.dataiq_button_stage2.grid(row=5, column=0, columnspan=3, pady=(0, 10))

        self.frame_stage2.columnconfigure(1, weight=1)
        self.frame_stage2.rowconfigure(2, weight=1)

        self.defined_column_groups = []
        self.all_loaded_headers = []
        self.header_checkbox_vars = []
        self.editing_group_index = None

        # --- Stage 3: Search Value in Excel Column ---
        self.frame_stage3 = LabelFrame(self.content_frame, text="Stage 3: Search Value in Excel Column", padx=20, pady=10)
        self.frame_stage3.pack(pady=10, padx=20, fill="x", expand=True)

        tk.Label(self.frame_stage3, text="Input Excel File (for Search):").grid(row=0, column=0, sticky="e", pady=5)
        self.input_search_excel_entry = tk.Entry(self.frame_stage3, width=40)
        self.input_search_excel_entry.grid(row=0, column=1, sticky="ew", pady=5)
        self.browse_search_excel_button = tk.Button(self.frame_stage3, text="Browse...", command=lambda: self.input_search_excel_entry.insert(0, filedialog.askopenfilename(title="Select Input Excel File")))
        self.browse_search_excel_button.grid(row=0, column=2, padx=5, pady=5)

        self.load_search_headers_button = tk.Button(self.frame_stage3, text="Load Columns", command=self.load_search_excel_columns)
        self.load_search_headers_button.grid(row=1, column=0, columnspan=3, pady=5)

        tk.Label(self.frame_stage3, text="Select Column:").grid(row=2, column=0, sticky="e", pady=5)
        self.search_column_combobox = ttk.Combobox(self.frame_stage3, width=37, state="disabled")
        self.search_column_combobox.grid(row=2, column=1, sticky="ew", pady=5)

        tk.Label(self.frame_stage3, text="Value to Search:").grid(row=3, column=0, sticky="e", pady=5)
        self.search_value_entry = tk.Entry(self.frame_stage3, width=40, state="disabled")
        self.search_value_entry.grid(row=3, column=1, sticky="ew", pady=5)

        self.search_button = tk.Button(
            self.frame_stage3,
            text="Search",
            width=20,
            bg="#17A2B8",
            fg="white",
            state="disabled",
            command=self.perform_search
        )
        self.search_button.grid(row=3, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage3, text="Search Results:").grid(row=4, column=0, sticky="nw", pady=5)
        self.search_results_frame = tk.Frame(self.frame_stage3)
        self.search_results_frame.grid(row=4, column=1, columnspan=2, sticky="nsew", pady=5)
        self.search_results_text = tk.Text(self.search_results_frame, height=10, width=60, state="disabled", wrap="none")
        self.search_results_text.pack(side="left", fill="both", expand=True)
        self.search_results_scrollbar = Scrollbar(self.search_results_frame, command=self.search_results_text.yview)
        self.search_results_scrollbar.pack(side="right", fill="y")
        self.search_results_text.config(yscrollcommand=self.search_results_scrollbar.set)

        self.frame_stage3.columnconfigure(1, weight=1)
        self.frame_stage3.rowconfigure(4, weight=1)

    # --- Stage 1 methods ---
    def select_input_text_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Input Text File (Stage 1)",
            filetypes=[("Text Files", "*.txt *.csv"), ("All Files", "*.*")]
        )
        if file_path:
            self.input_text_entry.delete(0, tk.END)
            self.input_text_entry.insert(0, file_path)

    def select_output_single_excel_file(self):
        file_path = filedialog.asksaveasfilename(
            title="Save Output Single Sheet Excel As (Stage 1)",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path:
            self.output_single_excel_entry.delete(0, tk.END)
            self.output_single_excel_entry.insert(0, file_path)

    def run_stage1_conversion_skip_rows(self):
        input_file = self.input_text_entry.get()
        output_file = self.output_single_excel_entry.get()
        delimiter = self.delimiter_entry.get()
        if not input_file:
            messagebox.showerror("Input Error", "Please select an Input Text File (Stage 1).")
            return
        if not output_file:
            messagebox.showerror("Output Error", "Please specify an Output Single Sheet Excel File (Stage 1).")
            return
        if not delimiter:
            messagebox.showerror("Input Error", "Please provide a Delimiter.")
            return
        self.convert_single_button.config(state=tk.DISABLED)
        self.convert_full_button.config(state=tk.DISABLED)
        self.root.update_idletasks()
        success, msg = self.convert_text_to_excel_skip_first_last(input_file, output_file, delimiter)
        self.convert_single_button.config(state=tk.NORMAL)
        self.convert_full_button.config(state=tk.NORMAL)
        if success:
            messagebox.showinfo("Stage 1 Success", msg)
        else:
            messagebox.showerror("Stage 1 Failed", msg)

    def run_stage1_conversion_full(self):
        input_file = self.input_text_entry.get()
        output_file = self.output_single_excel_entry.get()
        delimiter = self.delimiter_entry.get()
        if not input_file:
            messagebox.showerror("Input Error", "Please select an Input Text File (Stage 1).")
            return
        if not output_file:
            messagebox.showerror("Output Error", "Please specify an Output Single Sheet Excel File (Stage 1).")
            return
        if not delimiter:
            messagebox.showerror("Input Error", "Please provide a Delimiter.")
            return
        self.convert_single_button.config(state=tk.DISABLED)
        self.convert_full_button.config(state=tk.DISABLED)
        self.root.update_idletasks()
        success, msg = self.convert_text_to_excel_full(input_file, output_file, delimiter)
        self.convert_single_button.config(state=tk.NORMAL)
        self.convert_full_button.config(state=tk.NORMAL)
        if success:
            messagebox.showinfo("Stage 1 Success", msg)
        else:
            messagebox.showerror("Stage 1 Failed", msg)

    def convert_text_to_excel_skip_first_last(self, input_file, output_file, delimiter):
        try:
            with open(input_file, "r", encoding="utf-8") as f:
                lines = f.readlines()
            if len(lines) <= 2:
                return False, "Not enough lines in the text file."
            lines = lines[1:-1]
            rows = [line.strip().split(delimiter) for line in lines]
            rows = [[cell.strip('"') for cell in row] for row in rows]
            df = pd.DataFrame(rows)
            df.to_excel(output_file, index=False, header=False)
            return True, f"File converted and saved to {output_file}"
        except Exception as e:
            return False, str(e)

    def convert_text_to_excel_full(self, input_file, output_file, delimiter):
        try:
            with open(input_file, "r", encoding="utf-8") as f:
                lines = f.readlines()
            rows = [line.strip().split(delimiter) for line in lines]
            rows = [[cell.strip('"') for cell in row] for row in rows]
            df = pd.DataFrame(rows)
            df.to_excel(output_file, index=False, header=False)
            return True, f"File converted and saved to {output_file}"
        except Exception as e:
            return False, str(e)

    # --- Stage 2 methods (split, group management, scrollbars, etc.) ---
    def select_input_split_excel_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Input Excel File (Stage 2)",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path:
            self.input_split_excel_entry.delete(0, tk.END)
            self.input_split_excel_entry.insert(0, file_path)
            self.load_split_excel_headers()
            self.clear_defined_groups()
            self.hide_group_definition_frame()

    def select_output_split_folder(self):
        folder_path = filedialog.askdirectory(
            title="Select Output Folder for Split Files (Stage 2)"
        )
        if folder_path:
            self.output_split_folder_entry.delete(0, tk.END)
            self.output_split_folder_entry.insert(0, folder_path)

    def load_split_excel_headers(self):
        input_excel_file = self.input_split_excel_entry.get()
        self.all_loaded_headers = []
        self.add_group_button.config(state=tk.DISABLED)
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)
        self.clear_defined_groups()
        self.clear_header_checkboxes()
        self.hide_group_definition_frame()
        if not input_excel_file:
            return
        if not os.path.exists(input_excel_file):
            messagebox.showwarning("File Not Found", f"Input Excel file not found: {input_excel_file}")
            return
        try:
            headers_df = pd.read_excel(input_excel_file, sheet_name=0, nrows=0)
            headers = headers_df.columns.tolist()
            if headers:
                self.all_loaded_headers = headers
                self.create_header_checkboxes(headers)
                self.add_group_button.config(state=tk.NORMAL)
            else:
                messagebox.showwarning("No Headers Found", f"Could not detect headers in the first sheet of Excel file: {input_excel_file}.\nCheck if the first row contains headers.")
        except Exception as e:
            messagebox.showerror("Error Loading Excel Headers", str(e))

    def create_header_checkboxes(self, headers):
        self.clear_header_checkboxes()
        self.header_checkbox_vars = []
        for i, header in enumerate(headers):
            var = BooleanVar()
            cb = Checkbutton(self.headers_checkbox_frame, text=header, variable=var, anchor="w")
            cb.grid(row=i, column=0, sticky="w")
            self.header_checkbox_vars.append(var)
        self.headers_checkbox_frame.update_idletasks()
        self.on_headers_frame_configure(None)

    def clear_header_checkboxes(self):
        for widget in self.headers_checkbox_frame.winfo_children():
            widget.destroy()
        self.header_checkbox_vars = []

    def select_all_headers_checkboxes(self):
        for var in self.header_checkbox_vars:
            var.set(True)

    def deselect_all_headers_checkboxes(self):
        for var in self.header_checkbox_vars:
            var.set(False)

    def clear_defined_groups(self):
        self.defined_column_groups = []
        self.update_groups_listbox()
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)

    def update_groups_listbox(self):
        self.split_groups_listbox.delete(0, tk.END)
        if not self.defined_column_groups:
            self.split_groups_listbox.insert(tk.END, "No column groups defined yet.")
            self.edit_group_button.config(state=tk.DISABLED)
            self.remove_group_button.config(state=tk.DISABLED)
            self.perform_split_button.config(state=tk.DISABLED)
            return
        for i, (output_file_name, columns) in enumerate(self.defined_column_groups):
            display_text = f"Group {i+1}: {', '.join(columns)} -> {output_file_name}.xlsx"
            self.split_groups_listbox.insert(tk.END, display_text)
        self.perform_split_button.config(state=tk.NORMAL)
        self.split_groups_listbox.config(state=tk.NORMAL)

    def show_group_definition_frame(self):
        self.group_definition_frame.grid(row=6, column=0, columnspan=3, sticky="ew", pady=10)
        self.add_group_button.config(state=tk.DISABLED)
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)
        self.split_groups_listbox.config(state=tk.DISABLED)
        self.output_split_folder_entry.config(state=tk.DISABLED)
        self.browse_output_split_folder_button.config(state=tk.DISABLED)

    def hide_group_definition_frame(self):
        self.group_definition_frame.grid_forget()
        self.split_groups_listbox.config(state=tk.NORMAL)
        self.add_group_button.config(state=tk.NORMAL)
        self.output_split_folder_entry.config(state=tk.NORMAL)
        self.browse_output_split_folder_button.config(state=tk.NORMAL)
        self.on_group_select(None)

    def add_column_group(self):
        if not self.all_loaded_headers:
            messagebox.showwarning("Headers Not Loaded", "Please load headers from the Excel file first.")
            return
        self.editing_group_index = None
        self.output_file_name_entry.delete(0, tk.END)
        self.deselect_all_headers_checkboxes()
        self.group_definition_frame.config(text="Define New Column Group")
        self.show_group_definition_frame()

    def edit_selected_group(self):
        selected_indices = self.split_groups_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Selection Error", "Please select a column group from the list to edit.")
            return
        self.editing_group_index = selected_indices[0]
        output_file_name, columns = self.defined_column_groups[self.editing_group_index]
        self.output_file_name_entry.delete(0, tk.END)
        self.output_file_name_entry.insert(0, output_file_name)
        self.deselect_all_headers_checkboxes()
        for col in columns:
            try:
                index = self.all_loaded_headers.index(col)
                self.header_checkbox_vars[index].set(True)
            except ValueError:
                pass
        self.group_definition_frame.config(text=f"Edit Column Group {self.editing_group_index + 1}")
        self.show_group_definition_frame()

    def save_column_group(self):
        output_file_name = self.output_file_name_entry.get().strip()
        selected_columns = [self.all_loaded_headers[i] for i, var in enumerate(self.header_checkbox_vars) if var.get()]
        if not output_file_name:
            messagebox.showwarning("Input Error", "Please specify an output file name.")
            return
        output_file_name = re.sub(r'[^\w\s.-]', '', output_file_name)
        output_file_name = output_file_name.replace(' ', '_')
        if not selected_columns:
            messagebox.showwarning("Selection Error", "Please select at least one column for this group.")
            return
        if self.editing_group_index is None:
            self.defined_column_groups.append((output_file_name, selected_columns))
        else:
            self.defined_column_groups[self.editing_group_index] = (output_file_name, selected_columns)
        self.split_groups_listbox.config(state=tk.NORMAL)
        self.update_groups_listbox()
        self.hide_group_definition_frame()
        self.editing_group_index = None

    def cancel_column_group_edit(self):
        self.hide_group_definition_frame()
        self.editing_group_index = None

    def remove_selected_group(self):
        selected_indices = self.split_groups_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Selection Error", "Please select a column group from the list to remove.")
            return
        group_index = selected_indices[0]
        del self.defined_column_groups[group_index]
        self.update_groups_listbox()

    def perform_column_group_split(self):
        input_excel_file = self.input_split_excel_entry.get()
        output_folder = self.output_split_folder_entry.get()
        if not input_excel_file:
            messagebox.showerror("Input Error", "Please select an Input Excel File (Stage 2).")
            return
        if not output_folder:
            messagebox.showerror("Input Error", "Please specify an Output Folder (Stage 2).")
            return
        if not self.defined_column_groups:
            messagebox.showwarning("No Groups Defined", "Please define at least one column group to perform the split.")
            return
        if not os.path.exists(output_folder):
            try:
                os.makedirs(output_folder)
            except Exception as e:
                messagebox.showerror("Folder Creation Error", f"Could not create output folder: {e}")
                return
        output_file_names = [group[0] for group in self.defined_column_groups]
        column_groups_list = [group[1] for group in self.defined_column_groups]
        self.add_group_button.config(state=tk.DISABLED)
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)
        self.split_groups_listbox.config(state=tk.DISABLED)
        self.output_split_folder_entry.config(state=tk.DISABLED)
        self.browse_output_split_folder_button.config(state=tk.DISABLED)
        self.convert_single_button.config(state=tk.DISABLED)
        self.convert_full_button.config(state=tk.DISABLED)
        self.dataiq_button_stage2.config(state=tk.DISABLED)
        self.root.update_idletasks()
        try:
            df = pd.read_excel(input_excel_file, sheet_name=0, header=0, dtype=str)
            split_count = 0
            for output_file_name, columns_to_include in zip(output_file_names, column_groups_list):
                output_file_path = os.path.join(output_folder, f"{output_file_name}.xlsx")
                try:
                    missing_cols = [col for col in columns_to_include if col not in df.columns]
                    if missing_cols:
                        messagebox.showwarning("Missing Columns", f"Skipping group for '{output_file_name}.xlsx' due to missing columns in the first sheet: {', '.join(missing_cols)}")
                        continue
                    df_subset = df[columns_to_include]
                    df_subset.to_excel(output_file_path, index=False)
                    if os.path.exists(output_file_path) and os.path.getsize(output_file_path) > 100:
                        wb = load_workbook(output_file_path)
                        ws = wb.active
                        text_fmt = '@'
                        for row in ws.iter_rows():
                            for cell in row:
                                cell.number_format = text_fmt
                        wb.save(output_file_path)
                    split_count += 1
                except Exception as save_error:
                    messagebox.showwarning("Save Error", f"Could not save group to '{output_file_path}': {save_error}")
            if split_count > 0:
                messagebox.showinfo("Split Success", f"Successfully split Excel file into {split_count} files in folder: {output_folder}")
            else:
                messagebox.showwarning("Split Completed", f"Split operation completed, but no files were successfully created in folder: {output_folder}")
        except FileNotFoundError:
            messagebox.showerror("File Not Found", f"Input Excel file not found at {input_excel_file}")
        except Exception as e:
            messagebox.showerror("Split Failed", str(e))
        finally:
            self.convert_single_button.config(state=tk.NORMAL)
            self.convert_full_button.config(state=tk.NORMAL)
            self.dataiq_button_stage2.config(state=tk.NORMAL)
            self.input_split_excel_entry.config(state=tk.NORMAL)
            self.browse_split_excel_button.config(state=tk.NORMAL)
            self.output_split_folder_entry.config(state=tk.NORMAL)
            self.browse_output_split_folder_button.config(state=tk.NORMAL)
            if self.all_loaded_headers:
                self.add_group_button.config(state=tk.NORMAL)
            if self.defined_column_groups:
                self.perform_split_button.config(state=tk.NORMAL)
            self.split_groups_listbox.config(state=tk.NORMAL)

    # --- Stage 3 methods ---
    def load_search_excel_columns(self):
        input_excel_file = self.input_search_excel_entry.get()
        self.search_column_combobox.set('')
        self.search_column_combobox['values'] = []
        self.search_column_combobox.config(state="disabled")
        self.search_value_entry.delete(0, tk.END)
        self.search_value_entry.config(state="disabled")
        self.search_button.config(state="disabled")
        self.search_results_text.config(state="normal")
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.insert(tk.END, "Load an Excel file to see columns.")
        self.search_results_text.config(state="disabled")
        if not input_excel_file:
            messagebox.showwarning("Input Error", "Please select an Input Excel File for Search (Stage 3).")
            return
        if not os.path.exists(input_excel_file):
            messagebox.showwarning("File Not Found", f"Input Excel file not found: {input_excel_file}")
            return
        try:
            headers_df = pd.read_excel(input_excel_file, nrows=0)
            headers = headers_df.columns.tolist()
            if headers:
                self.search_column_combobox['values'] = headers
                self.search_column_combobox.config(state="readonly")
                self.search_results_text.config(state="normal")
                self.search_results_text.delete(1.0, tk.END)
                self.search_results_text.insert(tk.END, f"Columns loaded. Select a column and enter a value to search.")
                self.search_results_text.config(state="disabled")
                self.search_value_entry.config(state="normal")
                self.search_button.config(state="normal")
            else:
                messagebox.showwarning("No Headers Found", f"Could not detect headers in Excel file: {input_excel_file}.\nCheck if the first row contains headers.")
                self.search_results_text.config(state="normal")
                self.search_results_text.delete(1.0, tk.END)
                self.search_results_text.insert(tk.END, "No headers found in the selected file.")
                self.search_results_text.config(state="disabled")
        except Exception as e:
            messagebox.showerror("Error Loading Excel Headers", str(e))
            self.search_results_text.config(state="normal")
            self.search_results_text.delete(1.0, tk.END)
            self.search_results_text.insert(tk.END, f"An error occurred during the search: {e}")
            self.search_results_text.config(state="disabled")

    def perform_search(self):
        input_excel_file = self.input_search_excel_entry.get()
        selected_column = self.search_column_combobox.get()
        search_value = self.search_value_entry.get()
        self.search_results_text.config(state="normal")
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.config(state="disabled")
        if not input_excel_file:
            messagebox.showwarning("Input Error", "Please select an Input Excel File for Search (Stage 3).")
            return
        if not selected_column:
            messagebox.showwarning("Selection Error", "Please select a column to search in.")
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, "Please select a column.")
            self.search_results_text.config(state="disabled")
            return
        if not search_value:
            messagebox.showwarning("Input Error", "Please enter a value to search for.")
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, "Please enter a value to search.")
            self.search_results_text.config(state="disabled")
            return
        if not os.path.exists(input_excel_file):
            messagebox.showwarning("File Not Found", f"Input Excel file not found: {input_excel_file}")
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, "Error: File not found.")
            self.search_results_text.config(state="disabled")
            return
        try:
            df = pd.read_excel(input_excel_file, dtype=str)
            if selected_column not in df.columns:
                messagebox.showerror("Column Error", f"Selected column '{selected_column}' not found in the Excel file.")
                self.search_results_text.config(state="normal")
                self.search_results_text.insert(tk.END, f"Error: Column '{selected_column}' not found.")
                self.search_results_text.config(state="disabled")
                return
            matching_rows_df = df[df[selected_column].astype(str).str.contains(search_value, case=False, na=False, regex=False)]
            self.search_results_text.config(state="normal")
            self.search_results_text.delete(1.0, tk.END)
            if not matching_rows_df.empty:
                results_string = matching_rows_df.to_string(index=False)
                self.search_results_text.insert(tk.END, results_string)
            else:
                self.search_results_text.insert(tk.END, f"No results found for '{search_value}' in column '{selected_column}'.")
            self.search_results_text.config(state="disabled")
        except Exception as e:
            messagebox.showerror("Search Error", str(e))
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, f"An error occurred during the search: {e}")
            self.search_results_text.config(state="disabled")

    def on_group_select(self, event):
        if self.group_definition_frame.winfo_ismapped():
            self.edit_group_button.config(state=tk.DISABLED)
            self.remove_group_button.config(state=tk.DISABLED)
        else:
            if self.split_groups_listbox.curselection():
                self.edit_group_button.config(state=tk.NORMAL)
                self.remove_group_button.config(state=tk.NORMAL)
            else:
                self.edit_group_button.config(state=tk.DISABLED)
                self.remove_group_button.config(state=tk.DISABLED)

    def on_headers_frame_configure(self, event):
        self.headers_canvas.configure(scrollregion=self.headers_checkbox_frame.bbox("all"))

    def open_dataiq_url(self):
        dataiq_url = "https://www.example.com/dataiq"
        try:
            webbrowser.open(dataiq_url)
        except Exception as e:
            messagebox.showerror("Error Opening URL", f"Could not open the DataIQ URL: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelToolApp(root)
    root.mainloop()
